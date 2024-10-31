from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import getpass
import os
import csv
import sqlite3
import shutil
import time
import sys
import calendar
import pandas as pd
import mysql.connector
from mysql.connector import Error
from datetime import datetime, date
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.service import Service as ChromeService
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from time import sleep
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import Select
# from pynput.keyboard import Key, Controller
import re
from shareplum import Site
from shareplum.site import Version
import pyautogui 
import openpyxl


class Navegador:
    def __init__(self, usuario):
        self.options = self.config_options(usuario)
        self.navegador = webdriver.Chrome(options=self.options)
        self.wait = WebDriverWait(self.navegador, 50)
        self.connection = None  
        

    def config_options(self, usuario):
        download_directory = "C:\\Freightech\\"

        options = Options()
        options.add_argument('--disable-gpu')
        options.add_argument("--start-maximized")
        options.add_argument("--force-device-scale-factor=0.8")
        options.add_argument("--disable-javascript")
        
        
        # Configurações de preferências de download vai guardar nada 
        prefs = {
            "download.default_directory": download_directory,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        options.add_experimental_option("prefs", prefs)

        return options    
    

    def abrir_navegador(self, url):
        self.navegador.get(url)

    def limpar_campo(self, elemento):
        campo = self.wait.until( 
            EC.visibility_of_element_located((By.XPATH, elemento))
        )
        self.navegador.execute_script("arguments[0].value = '';", campo)
    
    @staticmethod
    def get_data_primeiro_dia_do_ano():
        data_atual = datetime.now()
        primeiro_dia_ano = data_atual.replace(month=1, day=1)     
        return primeiro_dia_ano.strftime("%d/%m/%Y")

    @staticmethod
    def get_data_primeiro_dia_do_mes():
        data_atual = datetime.now().replace(day=1)
        return data_atual.strftime("%d/%m/%Y")

    def wait(self, time):
        return WebDriverWait(self.navegador, time)

    def click(self, elemento):
        try:
            self.wait.until(EC.visibility_of_element_located((By.XPATH, elemento))).click()
        except Exception as e:
            print(f"Erro ao clicar no elemento: {e}")

    def excluir_arquivos_em_pasta(self, caminho_pasta):
        try:
            arquivos = os.listdir(caminho_pasta)
            for arquivo in arquivos:
                caminho_completo = os.path.join(caminho_pasta, arquivo)
                if os.path.isfile(caminho_completo):
                    os.remove(caminho_completo)
                    print(f'Arquivo excluído: {caminho_completo}')
            print('Todos os arquivos foram excluídos.')
        except Exception as e:
            print(f'Erro ao excluir arquivos: {e}')

    def escreva_digitando(self, elemento, valor):
        campo = self.wait.until(
            EC.visibility_of_element_located((By.XPATH, elemento))
        )
        campo.clear()
        for letra in valor:
            campo.send_keys(letra)
            time.sleep(0.1)
    
    # Removido o método escrever_digitando, pois era redundante
    # e causava o erro de atributo
            
    
    def hover_and_click(self, element_to_hover, element_to_click):
        try:
            action = ActionChains(self.navegador)
            element_hover = self.wait(50).until(EC.visibility_of_element_located((By.XPATH, element_to_hover)))
            element_click = self.wait(50).until(EC.visibility_of_element_located((By.XPATH, element_to_click)))
            action.move_to_element(element_hover).click(element_click).perform()
        except Exception as e:
            print(f"Erro ao realizar hover e clique: {e}")

    def entrarIframe(self, xpath):
        try:
            self.wait(50).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, xpath))
            )
        except Exception as e:
            print(f"Erro ao entrar no iframe: {e}")

    def sairIframe(self):
        try:
            self.navegador.switch_to.default_content()
        except Exception as e:
            print(f"Erro ao sair do iframe: {e}")

    def create_server_connection_mysql(self, host_name, user_name, user_password):
        try:
            self.connection = mysql.connector.connect(
                host=host_name,
                user=user_name,
                passwd=user_password
            )
            print("MySQL Database connection successful")
        except Error as err:
            print(f"Error: '{err}'")
            sys.exit()

    def read_query(self, query):
        if not self.connection:
            print("No MySQL connection established.")
            return None
        
        cursor = self.connection.cursor()
        result = None
        try:
            cursor.execute(query)
            result = cursor.fetchall()
            return result
        except Error as err:
            print(f"Error: '{err}'")
            sys.exit()
            
    def change_tab_with_interactions(self, url):
        try:
            # Salvar a janela atual 
            current_window = self.navegador.current_window_handle
            
            # Abrir uma nova aba e mudar para ela
            self.navegador.execute_script("window.open(arguments[0], '_blank');", url)
            
            # Mudar o foco para a nova aba
            for window_handle in self.navegador.window_handles:
                if window_handle != current_window:
                    self.navegador.switch_to.window(window_handle)
                    break
        except Exception as e:
            print(f"Erro ao mudar de aba: {e}")

  
    def selecione_primeiro_elemento(self, xpath_dropdown, xpath_opcoes):
        try:
            # XPath do dropdown
            xpath_dropdown = "/html/body/app-root/ft-home/div/main/div/wac-exportacao-dados/div[2]/wac-box/div/div/div[2]/ft-escolha-segmento/div/div[2]/div/wac-select/ng-select/ng-dropdown-panel"

            # Espera até que o dropdown esteja disponível e clicável
            self.wait.until(
                EC.element_to_be_clickable((By.XPATH, xpath_dropdown))
            ).click()
            time.sleep(1)  # Pequena espera para o dropdown expandir

            # Localiza todos os elementos dentro do dropdown
            xpath_opcoes = "//div[@id='a97bada09e2b-0']//div"
            opcoes = self.wait.until(
                EC.visibility_of_all_elements_located((By.XPATH, xpath_opcoes))
            )

            if opcoes:
                # Seleciona o primeiro elemento da lista
                primeiro_elemento = opcoes[0]

                # Rola a página até o primeiro elemento
                self.driver.execute_script("arguments[0].scrollIntoView();", primeiro_elemento)
                time.sleep(1)  # Aguarda um pouco após o scroll

                print(f"Selecionando a primeira opção: {primeiro_elemento.text}")
                primeiro_elemento.click()
            else:
                print("Nenhuma opção foi encontrada no dropdown.")
        except Exception as e:
            print(f"Erro: {e}")
            
    def scroll_down(self):
        try:
            body = self.navegador.find_element(By.TAG_NAME, 'body')  
            body.send_keys(Keys.ARROW_DOWN)  
            time.sleep(1)  
        except Exception as e:
            print(f"Erro ao rolar a página para baixo: {e}")
            
    def scroll_up(self):
        try:
            body = self.navegador.find_element(By.TAG_NAME, 'body')  # Usando self.navegador
            body.send_keys(Keys.ARROW_UP)  # Rolar para cima
            # time.sleep(1)  # Aguardar um segundo
        except Exception as e:
            print(f"Erro ao rolar a página para cima: {e}")        
            

    def fechar_primeira_janela(self):
        try:
            # Verifica se há janelas abertas 
            if len(self.navegador.window_handles) > 0:
                # Muda para a primeira janela
                self.navegador.switch_to.window(self.navegador.window_handles[0])
                # Fecha a janela atual
                self.navegador.close()
                
                # Se ainda houver janelas abertas, muda para a próxima
                if len(self.navegador.window_handles) > 0:
                    self.navegador.switch_to.window(self.navegador.window_handles[0])
            else:
                print("Não há janelas abertas para fechar.")
        except Exception as e:
            print(f"Erro ao fechar a janela: {e}")
            
    def fechar_primeira_janela(self):
        # Certificar-se de que existem janelas abertas
        if len(self.driver.window_handles) > 0:
            # Trocar para a primeira janela
            self.driver.switch_to.window(self.driver.window_handles[0])
            # Fechar a primeira janela
            self.driver.close()
            
            # Se houver mais de uma janela, trocar para a próxima janela disponível
            if len(self.driver.window_handles) > 0:
                self.driver.switch_to.window(self.driver.window_handles[0])
        else:
            print("Não há janelas abertas para fechar.")
    
    def mover_arquivo(self, tipo):
        origem = "C:\\Freightech\\"

        # Mapeamento dos tipos para os diretórios de destino
        destinos = {
            "rota": "C:\\Freightech\\load\\ROTA",
            "as": "C:\\Freightech\\load\\AS",
            "armazem": "C:\\Freightech\\load\\ARMAZEM",
            "empurrada": "C:\\Freightech\\load\\EMPURRADA",
            "insumos": "C:\\Freightech\\load\\INSUMOS"
        }

        # Verifica se o tipo é válido
        if tipo not in destinos:
            raise ValueError(f"Tipo desconhecido: {tipo}. Use 'rota', 'as', 'armazem', 'empurrada' ou 'insumos'.")

        endereco_destino = destinos[tipo]

        # Cria a pasta de destino se não existir
        os.makedirs(endereco_destino, exist_ok=True)
        print(f"Pasta de destino garantida: {endereco_destino}")

        # Remove o arquivo mais antigo na pasta de destino, se houver
        self.remover_arquivo_mais_antigo(endereco_destino)

        # Move os arquivos da pasta de origem
        self.mover_arquivos(origem, endereco_destino)

    def remover_arquivo_mais_antigo(self, pasta):
        arquivos = [os.path.join(pasta, f) for f in os.listdir(pasta) if f.endswith('.xlsx')]
        if arquivos:
            arquivo_mais_antigo = min(arquivos, key=os.path.getctime)
            os.remove(arquivo_mais_antigo)
            print(f"Arquivo mais antigo removido: {arquivo_mais_antigo}")

    def mover_arquivos(self, origem, destino):
        for arquivo in os.listdir(origem):
            if arquivo.endswith('.xlsx'):
                caminho_arquivo = os.path.join(origem, arquivo)
                data = self.extrair_data_do_nome(arquivo)
                if data:
                    novo_nome = f"{data}.xlsx"
                    novo_caminho = os.path.join(destino, novo_nome)
                    shutil.move(caminho_arquivo, novo_caminho)
                    return f"{novo_caminho}"
                else:
                    print(f"Data não encontrada no nome do arquivo: {arquivo}")

    def extrair_data_do_nome(self, nome_arquivo):
        nome_sem_extensao = os.path.splitext(nome_arquivo)[0]
        data_atual = datetime.now().strftime("%Y-%m-%d")
        hora_atual = datetime.now().strftime("%H-%M-%S")
        return f"{nome_sem_extensao}_{data_atual}_{hora_atual}"
     
    def refresh_pagina(self):
        try:
            self.navegador.refresh()
            print("Página recarregada com sucesso.")
        except Exception as e:
            print(f"Erro ao recarregar a página: {e}") 
     
     
     
class Validacao:  
    def __init__(self,caminho):
        self.caminho = caminho
        # self.conectar_banco()
        


    def verificar_integridade_xlsx_rota(self, caminho_arquivo):
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            
            # Verifica se há pelo menos uma aba no arquivo
            if len(workbook.sheetnames) == 0:
                return False

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Verifica se a aba está vazia
                if sheet.max_row == 0 or sheet.max_column == 0:
                    continue  # Pula para a próxima aba se esta estiver vazia
                
                # Verifica se há pelo menos uma linha com conteúdo
                conteudo_encontrado = False
                for row in sheet.iter_rows():
                    if any(cell.value is not None for cell in row):
                        conteudo_encontrado = True
                        break
                
                if not conteudo_encontrado:
                    continue  # Pula para a próxima aba se não houver conteúdo

                # Verifica se há pelo menos uma coluna com cabeçalho
                colunas_presentes = [cell.value for cell in sheet[1] if cell.value is not None]
                if len(colunas_presentes) == 0:
                    continue  # Pula para a próxima aba se não houver cabeçalhos

                # Se chegou até aqui, pelo menos uma aba válida foi encontrada
                return True
            
            # Se nenhuma aba válida foi encontrada
            return False
        
        except Exception as e:
            self.registrar_erro(f"Erro ao verificar integridade do arquivo: {e}")
            return False


    def verificar_integridade_xlsx_as(self, caminho_arquivo):
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
            
                if sheet.max_row == 0 or sheet.max_column == 0:
                    return False
                
                
                for row in sheet.iter_rows():
                    if all(cell.value is None for cell in row):
                        return False
                
                
                colunas_esperadas = ['Coluna1', 'Coluna2', 'Coluna3'] 
                colunas_presentes = [cell.value for cell in sheet[1]]
                if not all(coluna in colunas_presentes for coluna in colunas_esperadas):
                    return False
            
            return True
        
        except Exception as e:
            self.registrar_erro(f"Erro ao verificar integridade do arquivo: {e}")
            return False




    def verificar_integridade_xlsx_empurrada(self, caminho_arquivo):
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                
                if sheet.max_row == 0 or sheet.max_column == 0:
                    return False
                

                for row in sheet.iter_rows():
                    if all(cell.value is None for cell in row):
                        return False
                
                # Verifica se todas as colunas esperadas estão presentes
                colunas_esperadas = ['Coluna1', 'Coluna2', 'Coluna3']  
                colunas_presentes = [cell.value for cell in sheet[1]]
                if not all(coluna in colunas_presentes for coluna in colunas_esperadas):
                    return False
            
            return True
        
        except Exception as e:
            self.registrar_erro(f"Erro ao verificar integridade do arquivo: {e}")
            return False



    def verificar_integridade_xlsx_armazem(self, caminho_arquivo):
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
        
                if sheet.max_row == 0 or sheet.max_column == 0:
                    return False
                
             
                for row in sheet.iter_rows():
                    if all(cell.value is None for cell in row):
                        return False
                
             
                colunas_esperadas = ['Coluna1', 'Coluna2', 'Coluna3']  
                colunas_presentes = [cell.value for cell in sheet[1]]
                if not all(coluna in colunas_presentes for coluna in colunas_esperadas):
                    return False
            
            return True
        
        except Exception as e:
            self.registrar_erro(f"Erro ao verificar integridade do arquivo: {e}")
            return False




    def verificar_integridade_xlsx_insumos(self, caminho_arquivo):
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                
                if sheet.max_row == 0 or sheet.max_column == 0:
                    return False
                
                
                for row in sheet.iter_rows():
                    if all(cell.value is None for cell in row):
                        return False
                
                
                colunas_esperadas = ['Coluna1', 'Coluna2', 'Coluna3'] 
                colunas_presentes = [cell.value for cell in sheet[1]]
                if not all(coluna in colunas_presentes for coluna in colunas_esperadas):
                    return False
            
            return True
        
        except Exception as e:
            self.registrar_erro(f"Erro ao verificar integridade do arquivo: {e}")
            return False




    def registrar_erro(self, mensagem):
        # Método para registrar erros (implemente conforme necessário) 
        print(mensagem)        
        
        
    def conectar_banco(self, host_name, user_name, user_password,database):
        try:
            self.connection = mysql.connector.connect(
                host=host_name,
                user=user_name,
                password=user_password,
                database=database
            )
            print("Conexão ao banco de dados MySQL realizada com sucesso")
        except mysql.connector.Error as err:
            print(f"Erro ao conectar ao banco de dados: '{err}'")
        return self.connection
      
      
    